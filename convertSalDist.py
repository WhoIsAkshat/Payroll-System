import mysql.connector
import pandas as pd
import openpyxl
from tkinter import messagebox
from openpyxl.styles import Alignment, Font, Border, Side
import datetime
from datetime import date,timedelta
import calendar

def add_thick_outside_borders(ws, start_row, start_col, end_row, end_col):
    thick_side = Side(border_style="thick", color="000000")
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            border = Border()
            if cell.row == start_row:
                border = Border(top=thick_side)
            if cell.row == end_row:
                border = Border(bottom=thick_side)
            if cell.column == start_col:
                border = Border(left=thick_side)
            if cell.column == end_col:
                border = Border(right=thick_side)
            if cell.row == start_row and cell.column == start_col:
                border = Border(top=thick_side, left=thick_side)
            if cell.row == start_row and cell.column == end_col:
                border = Border(top=thick_side, right=thick_side)
            if cell.row == end_row and cell.column == start_col:
                border = Border(bottom=thick_side, left=thick_side)
            if cell.row == end_row and cell.column == end_col:
                border = Border(bottom=thick_side, right=thick_side)

            cell.border = border


def days_in_month(month_year):
    year, month = map(int, month_year.split('-'))
    num_days = calendar.monthrange(year, month)[1]
    return num_days

mydate = datetime.datetime.now()

## prev month
first_day_of_current_month = mydate.replace(day=1)
last_day_of_previous_month = first_day_of_current_month - timedelta(days=1)
previous_month = last_day_of_previous_month.strftime('%Y-%B')
prev_month = last_day_of_previous_month.strftime('%Y-%m')

month = '%'+ prev_month + '-%'
m_month = '%'+ prev_month + '%'

mymonth = last_day_of_previous_month.strftime('%B')
ddate = mydate.strftime("%d-%m-%Y")
tdays = days_in_month(prev_month)
file_name = "Salary_"+str(last_day_of_previous_month.strftime('%B-%Y'))+"_SalaryDist.xlsx"

def adjust_column_widths(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        if column == 'A':
            ws.column_dimensions[column].width = 3
            continue
        if column == 'B':
            ws.column_dimensions[column].width = 18
            continue
        if column == 'M':
            ws.column_dimensions[column].width = 10
            continue
        if column == 'N':
            ws.column_dimensions[column].width = 5
            continue
        if column == 'O':
            ws.column_dimensions[column].width = 5
            continue
        if column == 'P':
            ws.column_dimensions[column].width = 4
            continue
        if column == 'D':
            ws.column_dimensions[column].width = 3
            continue
        if column == 'H':
            ws.column_dimensions[column].width = 3
            continue
        if column == 'F':
            ws.column_dimensions[column].width = 3
            continue
        if column == 'J':
            ws.column_dimensions[column].width = 3
            continue
        if column == 'L':
            ws.column_dimensions[column].width = 10
            continue    
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width
    wb.save(filepath)

cnx = mysql.connector.connect(host='localhost', 
                                database='salary', 
                                user='root', 
                                password='123456')
cursor = cnx.cursor()
current_month_name = datetime.datetime.now().strftime("%B")
current_year = int(datetime.datetime.now().year)
query = f"SELECT * FROM pay WHERE month='{current_month_name}' and year={current_year}"
cursor.execute(query)
data_rows = cursor.fetchall()
cnx.close()

wb = openpyxl.Workbook()
ws = wb.active

ws.cell(row=2, column=1, value=f"Month:\t\t{mymonth}")
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
ws.merge_cells(start_row=2, start_column=6, end_row=2, end_column=7)
ws.merge_cells(start_row=2, start_column=8, end_row=2, end_column=9)
ws.cell(row=2, column=6, value="Paid On")
ws.cell(row=2, column=6).alignment = Alignment(horizontal='center')
ws.cell(row=2, column=8, value=f"{ddate}")
ws.cell(row=2, column=11, value=f"{tdays}")
ws.cell(row=2, column=8).alignment = Alignment(horizontal='center')
ws.cell(row=2, column=11).alignment = Alignment(horizontal='center')
ws.cell(row=2, column=12, value="Chq.")
ws.cell(row=2, column=13, value="Cash")
ws.cell(row=2, column=14, value="Days")
ws.cell(row=2, column=15, value="Hrs.")
ws.cell(row=2, column=16, value="T")
ws.cell(row=2, column=12).alignment = Alignment(horizontal='center')
ws.cell(row=2, column=13).alignment = Alignment(horizontal='center')
ws.cell(row=2, column=14).alignment = Alignment(horizontal='center')
ws.cell(row=2, column=15).alignment = Alignment(horizontal='center')
ws.cell(row=2, column=16).alignment = Alignment(horizontal='center')

add_thick_outside_borders(ws, 1, 1, 2, 16)

row = 3
for item in data_rows:
    try:
        cnx = mysql.connector.connect(host='localhost', 
                                database='salary', 
                                user='root', 
                                password='123456')
        cursor = cnx.cursor()
        query = f"SELECT Rem_Advance FROM emp WHERE emp_id={item[0]}"
        cursor.execute(query)
        rem_advance = cursor.fetchone()
        rem_advance = int(rem_advance[0])
        query = f"select SUM(amount) from advances where emp_id={item[0]} and type='CR' and day_date LIKE '{month}'"
        cursor.execute(query)
        taken_adv = cursor.fetchone()
        if taken_adv != (None,):
            taken_adv = int(taken_adv[0])
        else:
            taken_adv = 0
        query = f"select * from attendance_summary where emp_id={item[0]} and month LIKE '{m_month}'"
        cursor.execute(query)
        att = cursor.fetchone()
        cnx.close()
    except mysql.connector.Error as err:
            messagebox.showerror("Error", f"Something went wrong: {err}")

    ws.cell(row=row, column=1, value=item[0])
    ws.cell(row=row, column=1, value=item[0]).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=2, value=item[1])
    ws.cell(row=row, column=2, value=item[1]).alignment = Alignment(horizontal='left')
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=3, value=rem_advance - int(taken_adv) + item[6])
    ws.cell(row=row, column=4, value="+")
    ws.cell(row=row, column=5, value=taken_adv)
    ws.cell(row=row, column=6, value="=")
    ws.cell(row=row, column=7, value=rem_advance+ item[6])
    ws.cell(row=row, column=8, value="-")
    ws.cell(row=row, column=9, value=item[6])
    ws.cell(row=row, column=10, value="=")
    ws.cell(row=row, column=11, value=rem_advance)

    ws.cell(row=row, column=14, value=round(att[2]*tdays)/100)
    ws.cell(row=row, column=15, value=att[3])
    ws.cell(row=row, column=16, value=att[4]*8)

    ws.cell(row=row+1, column=2).font = Font(bold=True)
    ws.cell(row=row+1, column=3, value=item[10]+item[3]+item[4]+item[6])
    ws.cell(row=row+1, column=4, value="-")
    ws.cell(row=row+1, column=5, value=item[3]+item[4])
    ws.cell(row=row+1, column=6, value="-")
    ws.cell(row=row+1, column=7, value=item[6])
    ws.cell(row=row+1, column=8, value="=")
    ws.cell(row=row+1, column=9, value=item[10])
    ws.cell(row=row+1, column=9).font = Font(bold=True)
    ws.cell(row=row+1, column=12, value=item[13])
    ws.cell(row=row+1, column=13, value=item[14])

    add_thick_outside_borders(ws, row, 1, row+1, 16)

    row += 2 

wb.save(file_name)
adjust_column_widths(file_name)

def bold_range(filepath, range_string):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    cells = ws[range_string]
    for row in cells:
        for cell in row:
            cell.font = Font(bold=True)
    wb.save(filepath)

bold_range(file_name, "A2:P2")

def title_range(filepath):
    ss = str(last_day_of_previous_month.strftime('%B-%Y'))
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=16)
    ws.cell(row=1, column=1, value=f"Salary Slip:   {ss}")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=1, column=1).font = Font(bold=True)
    wb.save(filepath)

title_range(file_name)
